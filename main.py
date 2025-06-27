from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi.responses import StreamingResponse
from fastapi.responses import StreamingResponse
from fastapi import FastAPI, UploadFile, File
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import UploadFile, File
from openpyxl import load_workbook
from transformers import pipeline
from datetime import datetime
import pandas as pd
import io
import os

app = FastAPI()

# Cargo el modelo
modelo = pipeline(
    "text-classification",
    model="Dylan1012/modelo_roberta_postventa",
    top_k=None,
    device=0,
    truncation=True
)

@app.post("/clasificar")
async def clasificar_archivo(file: UploadFile = File(...)):
    try:
        
        df = pd.read_excel(file.file, engine='openpyxl')

        # Asegurar que las columnas necesarias estén presentes
        columnas_requeridas = ['Descripción', 'Acción Correctora', 'Fecha Creación', 'Creado por']
        if not all(col in df.columns for col in columnas_requeridas):
            return {"error": "Faltan columnas necesarias en el archivo."}

        # Convertir la columna de fecha a datetime
        df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación'], errors='coerce')

        # Filtrar por el día actual
        hoy = pd.Timestamp.today().normalize()
        ayer = hoy - pd.Timedelta(days=1)
        df_dia = df[df['Fecha Creación'].dt.normalize() == ayer].copy()

        if df_dia.empty:
            return {"mensaje": "No hay incidencias del día actual."}

        # Crear la entrada para el modelo
        df_dia['Descripción'] = df_dia['Descripción'].fillna('').astype(str)
        df_dia['Acción Correctora'] = df_dia['Acción Correctora'].fillna('').astype(str)
        
        sep_token = " [SEP] "
        
        df_dia['input_text'] = (
            df_dia['Descripción'] + sep_token +
            df_dia['Acción Correctora']
        )

        # Aplicar modelo (puede tardar)
        resultados = modelo(df_dia['input_text'].tolist())

        # Extraer categoría con mayor puntuación
        df_dia['Predicción'] = [max(r, key=lambda x: x['score'])['label'] for r in resultados]
        df_dia['Precisión'] = [max(r, key=lambda x: x['score'])['score'] for r in resultados]

        # Seleccionar solo las columnas deseadas
        columnas_salida = {
            'Número de caso': 'id_incidencia',
            'Creado por': 'Creado por',
            'ID_PR': 'ID_PR',
            'Descripción': 'Descripción',
            'Acción Correctora': 'Acción Correctora',
            'Predicción': 'Predicción',
            'Precisión': 'Precisión'
        }

        df_final = df_dia[list(columnas_salida.keys())].rename(columns=columnas_salida)

        # Agrupar por Predicción y crear DataFrame con separadores visuales
        bloques = []
        for categoria, grupo in df_final.groupby('Predicción'):
            bloques.append(grupo)
            bloques.append(pd.DataFrame([[""] * grupo.shape[1]], columns=grupo.columns))  # línea en blanco

        df_formateado = pd.concat(bloques, ignore_index=True)

        # Guardar sin estilos por ahora
        temp_output = io.BytesIO()
        df_formateado.to_excel(temp_output, index=False)
        temp_output.seek(0)

        # Cargar y aplicar estilos con openpyxl
        wb = load_workbook(temp_output)
        ws = wb.active # Aqui estoy seleccionando la hoja activa del excel

        # Estilo de cabecera
        header_fill = PatternFill(start_color="308ec9", end_color="308ec9", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Precisión":
                col_idx = idx
                break
            
        # Color de las filas alterno
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for cell in ws[row]:
                    cell.fill = PatternFill(start_color="cdcdcd", end_color="cdcdcd", fill_type="solid")
                    
        # Color de las filas según la precisión
        if col_idx is not None:
            col_letter = get_column_letter(col_idx)

            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0.9:
                        cell.fill = PatternFill(start_color="b9ff99", end_color="b9ff99", fill_type="solid")  # Verde
                        cell.font = Font(color="000000", bold=True)
                    elif cell.value >= 0.8:
                        cell.fill = PatternFill(start_color="ffec99", end_color="ffec99", fill_type="solid")  # Amarillo
                        cell.font = Font(color="000000", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="ff9999", end_color="ff9999", fill_type="solid")  # Rojo
                        cell.font = Font(color="000000", bold=True)
                        
        # Si la fila está vacía la relleno de color negro
        for row in range(2, ws.max_row + 1):
            if all(cell.value is None for cell in ws[row]):
                for cell in ws[row]:
                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Guardar con estilo
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # Responder al cliente
        return StreamingResponse(
            final_output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={f'Incidencias_{ayer.strftime('%Y-%m-%d')}'}.xlsx"}
        )

    except Exception as e:
        return (str("033[91m {ERROR:}\033[00m " + str(e)))
    